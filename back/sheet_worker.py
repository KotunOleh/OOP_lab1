from PySide6.QtWidgets import (QTabWidget, QTableWidget, QTableWidgetItem, 
                               QHeaderView, QMessageBox, QInputDialog, QMenu)
from PySide6.QtGui import QAction
from PySide6.QtCore import Qt, QPoint
from openpyxl.utils import get_column_letter

from utils.config import DEFAULT_ROWS, DEFAULT_COLS
from back.parser import (
    ASTNode, NumberNode, CellRefNode, RangeRefNode, BinaryOpNode, 
    FunctionNode, UnaryOpNode, ErrorNode, ParsingError,
    CircularReferenceError, ReferenceError
)
from back.calculator import FormulaCalculator


class SheetWorker:
    def __init__(self, tab_widget: QTabWidget, main_window):
        self.tab_widget = tab_widget
        self.main_window = main_window 
        self._setup_signals()

    def _setup_signals(self):
        self.tab_widget.currentChanged.connect(self.main_window.on_tab_changed)
        self.tab_widget.tabBar().setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.tab_widget.tabBar().customContextMenuRequested.connect(self.main_window.show_tab_context_menu)
        
    def get_current_table(self) -> QTableWidget | None:
        return self.tab_widget.currentWidget()

    def get_current_sheet_name(self) -> str | None:
        idx = self.tab_widget.currentIndex()
        if idx != -1:
            return self.tab_widget.tabText(idx)
        return None

    def clear_tabs(self):
        self.tab_widget.blockSignals(True)
        self.tab_widget.clear()
        self.tab_widget.blockSignals(False)

    def add_sheet_tab(self, sheet_name: str, sheet_data = None) -> QTableWidget:
        table_widget = self.create_new_table_widget()
        self.populate_table(table_widget, sheet_data)
        index = self.tab_widget.addTab(table_widget, sheet_name)
        self.tab_widget.setCurrentIndex(index)
        self.update_column_headers(table_widget)
        return table_widget
        
    def create_new_table_widget(self) -> QTableWidget:
        table_widget = QTableWidget()
        table_widget.setToolTip("Клацніть правою кнопкою миші для опцій")
        table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        table_widget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        
        table_widget.customContextMenuRequested.connect(self.main_window.show_context_menu)
        table_widget.itemChanged.connect(self.main_window.on_item_changed)
        table_widget.itemDoubleClicked.connect(self.main_window.on_item_double_clicked)
        return table_widget
    
    def add_new_sheet_action(self):
        if not self.main_window.current_workbook:
            QMessageBox.warning(self.main_window, "Помилка", "Спочатку створіть або відкрийте файл.")
            return

        sheet_name, ok = QInputDialog.getText(self.main_window, "Новий аркуш", "Введіть ім'я аркуша:")
        
        if ok and sheet_name:
            if sheet_name in self.main_window.current_workbook.sheetnames:
                QMessageBox.warning(self.main_window, "Помилка", "Аркуш з таким іменем вже існує.")
                return
            
            new_sheet = self.main_window.current_workbook.create_sheet(title=sheet_name)
            self.add_sheet_tab(sheet_name) 
            self.main_window.set_dirty(True)

    def populate_table(self, table_widget: QTableWidget, sheet) -> None:
        table_widget.blockSignals(True)
        
        max_row, max_col = DEFAULT_ROWS, DEFAULT_COLS
        if sheet:
            max_row = sheet.max_row
            max_col = sheet.max_column
            if max_row == 1 and max_col == 1 and sheet.cell(1, 1).value is None:
                 max_row, max_col = DEFAULT_ROWS, DEFAULT_COLS

        table_widget.setRowCount(max_row)
        table_widget.setColumnCount(max_col)
        self.update_column_headers(table_widget)

        if sheet:
            for row_idx, row in enumerate(sheet.iter_rows(max_row=max_row, max_col=max_col)):
                for col_idx, cell in enumerate(row):
                    cell_value = cell.value
                    cell_value_str = str(cell_value) if cell_value is not None else ""
                    item = QTableWidgetItem()
                    if cell_value_str.startswith("="):
                        item.setData(Qt.ItemDataRole.UserRole, cell_value_str)
                        item.setText("...") 
                    else:
                        item.setText(cell_value_str)
                    table_widget.setItem(row_idx, col_idx, item)
        else:
             for r in range(max_row):
                 for c in range(max_col):
                     table_widget.setItem(r, c, QTableWidgetItem(""))

        table_widget.blockSignals(False)
        
    def update_column_headers(self, table_widget: QTableWidget | None = None) -> None:
        if table_widget is None:
            table_widget = self.get_current_table()
        if not table_widget: return
        col_count = table_widget.columnCount()
        headers = [get_column_letter(i + 1) for i in range(col_count)]
        table_widget.setHorizontalHeaderLabels(headers)

    def add_row(self) -> None:
        table_widget = self.get_current_table()
        if not table_widget: return
        current_row = table_widget.rowCount()
        table_widget.insertRow(current_row)
        self.main_window.set_dirty(True)
        
    def add_column(self) -> None:
        table_widget = self.get_current_table()
        if not table_widget: return
        current_col = table_widget.columnCount()
        table_widget.insertColumn(current_col)
        self.update_column_headers(table_widget)
        self.main_window.set_dirty(True)

    def delete_row(self) -> None:
        table_widget = self.get_current_table()
        if not table_widget: return
        
        row_count = table_widget.rowCount()
        if row_count <= 0: return
            
        last_row_index = row_count - 1
        
        self.update_formulas_on_delete('row', last_row_index)
        table_widget.removeRow(last_row_index)
        self.main_window.set_dirty(True)
        self.main_window.recalculate_all_cells()

    def delete_column(self) -> None:
        table_widget = self.get_current_table()
        if not table_widget: return
        
        col_count = table_widget.columnCount()
        if col_count <= 0: return

        last_col_index = col_count - 1

        
        self.update_formulas_on_delete('col', last_col_index)
        table_widget.removeColumn(last_col_index)
        self.update_column_headers(table_widget)
        self.main_window.set_dirty(True)
        self.main_window.recalculate_all_cells()
            
    def update_formulas_on_delete(self, dimension: str, deleted_index: int):
        calculator = self.main_window.calculator 
        self.main_window.calculator.clear_caches()

        for tab_idx in range(self.tab_widget.count()):
            table = self.tab_widget.widget(tab_idx)
            for r in range(table.rowCount()):
                for c in range(table.columnCount()):
                    if dimension == 'row' and r == deleted_index: continue
                    if dimension == 'col' and c == deleted_index: continue
                        
                    item = table.item(r, c)
                    if not item: continue
                    
                    formula = item.data(Qt.ItemDataRole.UserRole)
                    if not formula or not formula.startswith("="):
                        continue
                        
                    try:
                        ast = calculator._get_ast(formula)
                        new_ast = self._transform_ast_on_delete(ast, dimension, deleted_index, calculator)
                        
                        if not isinstance(new_ast, ErrorNode):
                            new_ast = self._check_bounds_after_delete(new_ast, dimension, table, calculator)
                        
                        if new_ast is ast:
                            continue

                        new_formula = "=" + new_ast.to_string()
                        item.setData(Qt.ItemDataRole.UserRole, new_formula)
                        
                        if self.main_window.is_formula_view:
                            item.setText(new_formula)
                        else:
                            item.setText("#REF!")
                            
                    except (ParsingError, ReferenceError, CircularReferenceError):
                        continue 
    
    def _check_bounds_after_delete(self, node: ASTNode, dim: str, table: QTableWidget, calc: FormulaCalculator) -> ASTNode:
        
        if isinstance(node, (NumberNode, ErrorNode)):
            return node
        
        if isinstance(node, CellRefNode):
            indices = calc.cell_name_to_indices(node.cell_name)
            if not indices: 
                return node
            r, c = indices
            if r >= table.rowCount() or c >= table.columnCount():
                return ErrorNode("#REF!")
            return node
        
        if isinstance(node, RangeRefNode):
            start_idx = calc.cell_name_to_indices(node.start_cell)
            end_idx = calc.cell_name_to_indices(node.end_cell)
            if not start_idx or not end_idx:
                return node
            r1, c1 = start_idx
            r2, c2 = end_idx
            if r1 >= table.rowCount() or c1 >= table.columnCount() or r2 >= table.rowCount() or c2 >= table.columnCount():
                return ErrorNode("#REF!")
            return node
        
        if isinstance(node, UnaryOpNode):
            return UnaryOpNode(node.op, self._check_bounds_after_delete(node.operand, dim, table, calc))
        
        if isinstance(node, BinaryOpNode):
            return BinaryOpNode(
                self._check_bounds_after_delete(node.left, dim, table, calc),
                node.op,
                self._check_bounds_after_delete(node.right, dim, table, calc)
            )
        
        if isinstance(node, FunctionNode):
            new_args = [self._check_bounds_after_delete(arg, dim, table, calc) for arg in node.args]
            return FunctionNode(node.func_name, new_args)
        
        return node 
                        
    def _transform_ast_on_delete(self, node: ASTNode, dim: str, idx: int, calc: FormulaCalculator) -> ASTNode:
        
        if isinstance(node, (NumberNode, ErrorNode)):
            return node 
        
        if isinstance(node, CellRefNode):
            indices = calc.cell_name_to_indices(node.cell_name)
            if not indices: return ErrorNode("#NAME?")
            
            r, c = indices
            
            if dim == 'row' and r == idx: return ErrorNode("#REF!") 
            if dim == 'col' and c == idx: return ErrorNode("#REF!")
            
            return node

        if isinstance(node, RangeRefNode):
            start_indices = calc.cell_name_to_indices(node.start_cell)
            end_indices = calc.cell_name_to_indices(node.end_cell)
            if not start_indices or not end_indices: return ErrorNode("#NAME?")

            r1, c1 = start_indices
            r2, c2 = end_indices
            
            if dim == 'row' and (r1 <= idx <= r2 or r2 <= idx <= r1):
                return ErrorNode("#REF!")
            if dim == 'col' and (c1 <= idx <= c2 or c2 <= idx <= c1):
                return ErrorNode("#REF!")

            return node

        if isinstance(node, UnaryOpNode):
            return UnaryOpNode(node.op, self._transform_ast_on_delete(node.operand, dim, idx, calc))

        if isinstance(node, BinaryOpNode):
            return BinaryOpNode(
                self._transform_ast_on_delete(node.left, dim, idx, calc),
                node.op,
                self._transform_ast_on_delete(node.right, dim, idx, calc)
            )

        if isinstance(node, FunctionNode):
            new_args = [self._transform_ast_on_delete(arg, dim, idx, calc) for arg in node.args]
            return FunctionNode(node.func_name, new_args)

        return node 
    
    def update_sheet_from_table(self, sheet, table_widget):
        if sheet.max_row > 0:
            sheet.delete_rows(1, sheet.max_row)
        
        for r in range(table_widget.rowCount()):
            for c in range(table_widget.columnCount()):
                item = table_widget.item(r, c)
                value_to_save = None 
                if item:
                    formula = item.data(Qt.ItemDataRole.UserRole)
                    if formula:
                        value_to_save = formula
                    else:
                        value_to_save = item.text()
                
                if value_to_save:
                    sheet.cell(row=r + 1, column=c + 1, value=value_to_save)

    def update_workbook_from_all_tabs(self, workbook):
        if not workbook: return
        for idx in range(self.tab_widget.count()):
            sheet_name = self.tab_widget.tabText(idx)
            if sheet_name not in workbook.sheetnames:
                continue 
            sheet = workbook[sheet_name]
            table_widget = self.tab_widget.widget(idx)
            self.update_sheet_from_table(sheet, table_widget)

    def populate_all_tabs(self, workbook) -> None:
        self.clear_tabs()
        
        self.tab_widget.blockSignals(True)
        try:
            if not workbook.sheetnames:
                self.add_sheet_tab("Sheet1")
            else:
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    self.add_sheet_tab(sheet_name, sheet)
            
            self.tab_widget.setCurrentIndex(0)
        finally:
            self.tab_widget.blockSignals(False)
        self.main_window.recalculate_all_cells()