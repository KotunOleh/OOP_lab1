import re
from openpyxl.utils import column_index_from_string, get_column_letter
from PySide6.QtCore import Qt
from back.parser import Parser, ErrorNode, ParsingError, CircularReferenceError, ReferenceError, ASTNode, NumberNode, CellRefNode, RangeRefNode, BinaryOpNode, FunctionNode, UnaryOpNode

class FormulaCalculator:
    def __init__(self):
        self._cell_name_cache = {}
        self._ast_cache: dict[str, object] = {}

    def clear_caches(self) -> None:
        try:
            self._cell_name_cache.clear()
        except Exception:
            pass
        try:
            self._ast_cache.clear()
        except Exception:
            pass

    def _get_ast(self, formula_string: str):
        if not isinstance(formula_string, str):
            return ErrorNode("#ERROR!")

        cached = self._ast_cache.get(formula_string)
        if cached is not None:
            return cached

        try:
            parser = Parser()
            ast = parser.parse(formula_string)
        except ParsingError:
            ast = ErrorNode("#ERROR!")
        except Exception:
            ast = ErrorNode("#ERROR!")

        try:
            self._ast_cache[formula_string] = ast
        except Exception:
            pass
        return ast

    @staticmethod
    def _excel_max(*args) -> float | int:
        if not args: return 0
        try: return max(args)
        except TypeError: return 0 

    @staticmethod
    def _excel_min(*args) -> float | int:
        if not args: return 0
        try: return min(args)
        except TypeError: return 0

    def _evaluate_ast(self, node: ASTNode, table_widget: any, visited: set[str] | None = None):
        if visited is None:
            visited = set()

        if isinstance(node, NumberNode):
            return node.value
        
        if isinstance(node, ErrorNode):
            raise ReferenceError(node.error_code)

        if isinstance(node, CellRefNode):
            cell = node.cell_name.upper()
            if cell in visited:
                raise CircularReferenceError(f"Circular reference detected at {cell}")
            visited.add(cell)

            indices = self.cell_name_to_indices(cell)
            if not indices:
                visited.discard(cell)
                raise ReferenceError("#NAME?")
            r, c = indices
            # If referenced cell is outside the current table bounds -> REF error
            if r >= table_widget.rowCount() or c >= table_widget.columnCount():
                visited.discard(cell)
                raise ReferenceError("#REF!")

            item = table_widget.item(r, c)
            if not item or not item.text():
                visited.discard(cell)
                return 0.0

            formula = item.data(Qt.ItemDataRole.UserRole) if hasattr(item, 'data') else None
            if formula and isinstance(formula, str) and formula.startswith("="):
                ast = self._get_ast(formula)
                value = self._evaluate_ast(ast, table_widget, visited)
                visited.discard(cell)
                return value

            try:
                val = float(item.text())
            except Exception:
                val = 0.0
            visited.discard(cell)
            return val

        if isinstance(node, RangeRefNode):
            start = node.start_cell
            end = node.end_cell
            start_idx = self.cell_name_to_indices(start)
            end_idx = self.cell_name_to_indices(end)
            if not start_idx or not end_idx:
                raise ReferenceError("#NAME?")
            r1, c1 = start_idx
            r2, c2 = end_idx
            min_r, max_r = min(r1, r2), max(r1, r2)
            min_c, max_c = min(c1, c2), max(c1, c2)
            values = []
            for rr in range(min_r, max_r + 1):
                for cc in range(min_c, max_c + 1):
                    cell_name = get_column_letter(cc + 1) + str(rr + 1)
                    try:
                        val = self._evaluate_ast(CellRefNode(cell_name), table_widget, visited)
                    except ReferenceError as e:
                        msg = str(e)
                        if msg.startswith("#REF"):
                            raise
                        continue
                    values.append(val)
            return values

        # Unary op
        if isinstance(node, UnaryOpNode):
            val = self._evaluate_ast(node.operand, table_widget, visited)
            if node.op == '-':
                return -val
            return val

        # Binary op
        if isinstance(node, BinaryOpNode):
            left = self._evaluate_ast(node.left, table_widget, visited)
            right = self._evaluate_ast(node.right, table_widget, visited)
            op = node.op
            try:
                if op == '+':
                    return left + right
                if op == '-':
                    return left - right
                if op == '*':
                    return left * right
                if op == '/':
                    if right == 0:
                        raise ReferenceError("#DIV/0!")
                    return left / right
                if op == '^':
                    if left == 0 and right == 0:
                        raise ReferenceError("#NUM!")
                    return left ** right
            except ReferenceError:
                raise
            except Exception:
                raise ReferenceError("#ERROR!")

        if isinstance(node, FunctionNode):
            func = node.func_name.upper()
            args_values = []
            for arg in node.args:
                val = self._evaluate_ast(arg, table_widget, visited)
                if isinstance(val, list):
                    args_values.extend(val)
                else:
                    args_values.append(val)

            nums = [v for v in args_values if isinstance(v, (int, float))]
            try:
                if func == 'SUM':
                    return sum(nums)
                if func == 'AVERAGE':
                    return sum(nums) / len(nums) if nums else 0
                if func == 'MAX':
                    return max(nums) if nums else 0
                if func == 'MIN':
                    return min(nums) if nums else 0
            except Exception:
                raise ReferenceError("#ERROR!")

        raise ReferenceError("#ERROR!")
    def cell_name_to_indices(self, cell_name: str) -> tuple[int, int] | None:
        if cell_name in self._cell_name_cache:
            return self._cell_name_cache[cell_name]
        match = re.match(r"([A-Z]+)([0-9]+)", cell_name.upper())
        if not match: return None
        col_str, row_str = match.groups()
        try:
            col_idx = column_index_from_string(col_str) - 1
            row_idx = int(row_str) - 1
            result = (row_idx, col_idx)
            self._cell_name_cache[cell_name] = result
            return result
        except Exception:
            return None

    def get_cell_value(self, cell_name: str, table_widget: any) -> float:
        indices = self.cell_name_to_indices(cell_name)
        if indices is None: return 0.0
        row, col = indices
        if row >= table_widget.rowCount() or col >= table_widget.columnCount():
            return 0.0
        item = table_widget.item(row, col)
        if not item or not item.text(): return 0.0
        try:
            return float(item.text())
        except ValueError:
            return 0.0


    def get_range_values(self, range_str: str, table_widget: any) -> list[float]:
        try:
            start_cell, end_cell = range_str.split(':')
            start_indices = self.cell_name_to_indices(start_cell)
            end_indices = self.cell_name_to_indices(end_cell)
            if start_indices is None or end_indices is None: return []
            r1, c1 = start_indices
            r2, c2 = end_indices
            min_row, max_row = min(r1, r2), max(r1, r2)
            min_col, max_col = min(c1, c2), max(c1, c2)
            values = []
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    cell_name = get_column_letter(c + 1) + str(r + 1)
                    values.append(self.get_cell_value(cell_name, table_widget))
            return [v for v in values if isinstance(v, (int, float))]
        except Exception:
            return []


    def parse_and_calculate(self, formula_string: str, table_widget: any) -> str:
        if not formula_string.startswith("="):
            return formula_string
        try:
            ast = self._get_ast(formula_string)
            if isinstance(ast, ErrorNode):
                return ast.error_code if hasattr(ast, 'error_code') else "#ERROR!"

            try:
                value = self._evaluate_ast(ast, table_widget)
            except CircularReferenceError:
                return "#CIRCULAR!"
            except ReferenceError as re_err:
                return str(re_err)
            if isinstance(value, list):
                return str(value)
            return str(value)
        except Exception:
            return "#ERROR!"