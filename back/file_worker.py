import openpyxl
import io
from openpyxl.workbook import Workbook
from PySide6.QtWidgets import QFileDialog, QMessageBox

from utils.config import DEFAULT_SHEET_NAME

class FileWorker:
    def __init__(self, parent_window):
        self.parent = parent_window

    def create_new_workbook(self) -> Workbook | None:
        try:
            workbook = Workbook()
            if len(workbook.sheetnames) > 0 and workbook.sheetnames[0] == 'Sheet':
                 if len(workbook.sheetnames) == 1:
                      workbook.active.title = DEFAULT_SHEET_NAME
                 else:
                     workbook.remove(workbook['Sheet'])

            if not workbook.sheetnames:
                 workbook.create_sheet(title=DEFAULT_SHEET_NAME)

            return workbook
        except Exception as e:
            QMessageBox.critical(self.parent, "Помилка", f"Не вдалося створити нову книгу: {e}")
            return None

    def open_local_workbook(self) -> tuple[Workbook | None, str | None]:
        options = QFileDialog.Options()
        filePath, _ = QFileDialog.getOpenFileName(
            self.parent, "Відкрити Excel файл", "",
            "Excel Files (*.xlsx);;All Files (*)", options=options
        )
        if filePath:
            try:
                workbook = openpyxl.load_workbook(filePath, data_only=False)
                return workbook, filePath
            except Exception as e:
                QMessageBox.critical(self.parent, "Помилка", f"Не вдалося відкрити файл: {e}")
        return None, None

    def save_local_workbook(self, workbook: Workbook, current_path: str | None) -> tuple[bool, str | None]:
        savePath = current_path
        if savePath is None:
            options = QFileDialog.Options()
            savePath, _ = QFileDialog.getSaveFileName(
                self.parent, "Зберегти новий файл", "Untitled.xlsx", 
                "Excel Files (*.xlsx);;All Files (*)", options=options
            )
        
        if savePath:
            try:
                workbook.save(savePath)
                QMessageBox.information(self.parent, "Успіх", f"Файл успішно збережено у:\n{savePath}")
                return True, savePath
            except Exception as e:
                QMessageBox.critical(self.parent, "Помилка", f"Не вдалося зберегти файл: {e}")
        return False, current_path

    def save_workbook_to_buffer(self, workbook: Workbook) -> io.BytesIO | None:
        """Зберігає книгу в BytesIO для передачі, наприклад, на Google Drive."""
        try:
            buffer = io.BytesIO()
            workbook.save(buffer)
            buffer.seek(0)
            return buffer
        except Exception as e:
            QMessageBox.critical(self.parent, "Помилка", f"Не вдалося зберегти книгу в буфер: {e}")
            return None

    def load_workbook_from_buffer(self, buffer: io.BytesIO) -> Workbook | None:
         """Завантажує книгу з BytesIO."""
         try:
              workbook = openpyxl.load_workbook(buffer, data_only=False)
              return workbook
         except Exception as e:
              QMessageBox.critical(self.parent, "Помилка", f"Не вдалося завантажити книгу з буфера: {e}")
              return None