import sys
import openpyxl
import re 
from openpyxl.workbook import Workbook
from openpyxl.utils import column_index_from_string, get_column_letter 

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout,
    QTableWidget, QTableWidgetItem, QFileDialog,
    QMessageBox, QHeaderView, QToolBar, QStyle, QInputDialog,
    QMenu
)
from PySide6.QtGui import QAction, QIcon
from PySide6.QtCore import Qt, QTimer, QPoint


class FormulaCalculator:
    """Обробляє парсинг та обчислення базових формул."""
    
    def __init__(self):
        self._cell_name_cache = {}

    @staticmethod
    def _excel_max(*args):
        if not args: return 0
        try: return max(args)
        except TypeError: return 0 

    @staticmethod
    def _excel_min(*args):
        if not args: return 0
        try: return min(args)
        except TypeError: return 0

    def cell_name_to_indices(self, cell_name):
        if cell_name in self._cell_name_cache:
            return self._cell_name_cache[cell_name]
        match = re.match(r"([A-Z]+)([0-9]+)", cell_name.upper())
        if not match: return None
        col_str, row_str = match.groups()
        try:
            col_idx = column_index_from_string(col_str) - 1
            row_idx = int(row_str) - 1
            result = (row_idx, col_idx)
            self._cell_name_cache[cell_name] = result
            return result
        except Exception:
            return None

    def get_cell_value(self, cell_name, table_widget):
        indices = self.cell_name_to_indices(cell_name)
        if indices is None: return 0.0
        row, col = indices
        if row >= table_widget.rowCount() or col >= table_widget.columnCount():
            return 0.0
        item = table_widget.item(row, col)
        if not item or not item.text(): return 0.0
        try:
            return float(item.text())
        except ValueError:
            return 0.0

    def get_range_values(self, range_str, table_widget):
        try:
            start_cell, end_cell = range_str.split(':')
            start_indices = self.cell_name_to_indices(start_cell)
            end_indices = self.cell_name_to_indices(end_cell)
            if start_indices is None or end_indices is None: return []
            r1, c1 = start_indices
            r2, c2 = end_indices
            min_row, max_row = min(r1, r2), max(r1, r2)
            min_col, max_col = min(c1, c2), max(c1, c2)
            values = []
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    cell_name = get_column_letter(c + 1) + str(r + 1)
                    values.append(self.get_cell_value(cell_name, table_widget))
            return [v for v in values if isinstance(v, (int, float))]
        except Exception:
            return []

    def parse_and_calculate(self, formula_string, table_widget):
        if not formula_string.startswith("="):
            return formula_string
        expression = formula_string[1:].upper()
        try:
            # SUM()
            match = re.match(r"SUM\((.*?)\)", expression)
            if match:
                range_str = match.group(1)
                if ":" in range_str:
                    values = self.get_range_values(range_str, table_widget)
                    return str(sum(values))
            # AVERAGE()
            match = re.match(r"AVERAGE\((.*?)\)", expression)
            if match:
                range_str = match.group(1)
                if ":" in range_str:
                    values = self.get_range_values(range_str, table_widget)
                    if not values: return "0"
                    return str(sum(values) / len(values))
            # MAX()
            match = re.match(r"MAX\((.*?)\)", expression)
            if match:
                range_str = match.group(1)
                if ":" in range_str:
                    values = self.get_range_values(range_str, table_widget)
                    if not values: return "0"
                    return str(max(values))
            # MIN()
            match = re.match(r"MIN\((.*?)\)", expression)
            if match:
                range_str = match.group(1)
                if ":" in range_str:
                    values = self.get_range_values(range_str, table_widget)
                    if not values: return "0"
                    return str(min(values))

            cell_names = re.findall(r"([A-Z]+[0-9]+)", expression)
            eval_expression = expression
            for cell_name in set(cell_names): 
                value = self.get_cell_value(cell_name, table_widget)
                eval_expression = re.sub(r"\b" + cell_name + r"\b", str(value), eval_expression)
            
            eval_expression = eval_expression.replace("^", "**")
            eval_expression = eval_expression.replace("=", "==")
            eval_expression = eval_expression.replace("!==", "!=")
            eval_expression = eval_expression.replace("<==", "<=")
            eval_expression = eval_expression.replace(">==", ">=")
            eval_expression = eval_expression.replace("<>", "!=")
            
            eval_globals = {}
            eval_locals = {
                "MAX": FormulaCalculator._excel_max,
                "MIN": FormulaCalculator._excel_min,
            }
            result = eval(eval_expression, eval_globals, eval_locals)
            return str(result)
        except Exception as e:
            return "#ERROR!"



class ExcelEditor(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("KotunSpreadSheeter")
        self.setGeometry(100, 100, 800, 600)

        self.current_workbook = None
        self.current_filepath = None
        self.active_sheet = None

        self.calculator = FormulaCalculator()
        self.is_calculating = False 
        
        self.is_formula_view = False

        # TOOLBAR
        toolbar = QToolBar("Головна панель")
        toolbar.setMovable(False)
        self.addToolBar(toolbar)
        
        style = self.style()
        # ICONS
        icon_new = style.standardIcon(QStyle.StandardPixmap.SP_FileIcon)
        icon_open = style.standardIcon(QStyle.StandardPixmap.SP_DialogOpenButton)
        icon_save = style.standardIcon(QStyle.StandardPixmap.SP_DialogSaveButton)
        icon_formulas = style.standardIcon(QStyle.StandardPixmap.SP_FileDialogDetailedView)

        # ACTIONS
        self.action_new = QAction(icon_new, "Створити", self)
        self.action_new.setToolTip("Створити новий файл (Ctrl + N)")
        self.action_new.setShortcut("Ctrl + N")
        self.action_new.triggered.connect(self.new_file)
        toolbar.addAction(self.action_new)

        self.action_open = QAction(icon_open, "Відкрити", self)
        self.action_open.setToolTip("Відкрити файл (Ctrl + O)")
        self.action_open.setShortcut("Ctrl + O")
        self.action_open.triggered.connect(self.open_file)
        toolbar.addAction(self.action_open)

        self.action_save = QAction(icon_save, "Зберегти", self)
        self.action_save.setToolTip("Зберегти файл (Ctrl + S)")
        self.action_save.setShortcut("Ctrl + S")
        self.action_save.triggered.connect(self.save_file)
        self.action_save.setEnabled(False)
        toolbar.addAction(self.action_save)

        toolbar.addSeparator()

        self.action_show_formulas = QAction(icon_formulas, "Показати формули", self)
        self.action_show_formulas.setToolTip("Режим перегляду формул (Ctrl + F)")
        self.action_show_formulas.setShortcut("Ctrl + F")
        self.action_show_formulas.setCheckable(True)
        self.action_show_formulas.toggled.connect(self.toggle_formula_view)
        self.action_show_formulas.setEnabled(False)
        toolbar.addAction(self.action_show_formulas)

        # CONTEXT MENU
        self.action_add_row = QAction("Додати рядок", self)
        self.action_add_row.triggered.connect(self.add_row)
        self.action_delete_row = QAction("Видалити поточний рядок", self)
        self.action_delete_row.triggered.connect(self.delete_row)
        self.action_add_column = QAction("Додати стовпець", self)
        self.action_add_column.triggered.connect(self.add_column)
        self.action_delete_column = QAction("Видалити поточний стовпець", self)
        self.action_delete_column.triggered.connect(self.delete_column)
        self.set_edit_actions_enabled(False)

        # TABLE
        self.table_widget = QTableWidget()
        self.table_widget.setToolTip("Клацніть правою кнопкою миші для опцій")
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_widget.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_widget.customContextMenuRequested.connect(self.show_context_menu)
        self.table_widget.itemChanged.connect(self.on_item_changed)
        self.table_widget.itemDoubleClicked.connect(self.on_item_double_clicked)
        self.setCentralWidget(self.table_widget)

    
    def show_context_menu(self, position: QPoint) -> None:
        if not self.action_save.isEnabled(): return
        context_menu = QMenu(self)
        current_row = self.table_widget.currentRow()
        current_col = self.table_widget.currentColumn()
        self.action_delete_row.setEnabled(current_row >= 0)
        self.action_delete_column.setEnabled(current_col >= 0)
        context_menu.addAction(self.action_add_row)
        context_menu.addAction(self.action_delete_row)
        context_menu.addSeparator()
        context_menu.addAction(self.action_add_column)
        context_menu.addAction(self.action_delete_column)
        global_pos = self.table_widget.mapToGlobal(position)
        context_menu.exec(global_pos)

    def add_row(self) -> None:
        current_row = self.table_widget.currentRow()
        if current_row < 0: current_row = self.table_widget.rowCount()
        self.table_widget.insertRow(current_row)

    def delete_row(self) -> None:
        current_row = self.table_widget.currentRow()
        if current_row >= 0: self.table_widget.removeRow(current_row)

    def add_column(self) -> None:
        current_col = self.table_widget.currentColumn()
        if current_col < 0: current_col = self.table_widget.columnCount()
        self.table_widget.insertColumn(current_col)
        self.update_column_headers()

    def delete_column(self) -> None:
        current_col = self.table_widget.currentColumn()
        if current_col >= 0:
            self.table_widget.removeColumn(current_col)
            self.update_column_headers()

    def update_column_headers(self) -> None:
        col_count = self.table_widget.columnCount()
        headers = [get_column_letter(i + 1) for i in range(col_count)]
        self.table_widget.setHorizontalHeaderLabels(headers)

    def set_edit_actions_enabled(self, enabled: bool) -> None:
        self.action_add_row.setEnabled(enabled)
        self.action_delete_row.setEnabled(enabled)
        self.action_add_column.setEnabled(enabled)
        self.action_delete_column.setEnabled(enabled)

    
    # UTILS
    def toggle_formula_view(self, checked: bool) -> None:
        """Switch formula view modes"""
        self.is_formula_view = checked
        self.recalculate_all_cells()

    def on_item_double_clicked(self, item: QTableWidgetItem) -> None:
        """Коли користувач двічі клікає, ЗАВЖДИ показати йому ФОРМУЛУ."""
        if self.is_calculating: return
        formula = item.data(Qt.ItemDataRole.UserRole)
        if formula:
            self.is_calculating = True 
            item.setText(formula)    
            self.is_calculating = False

    def on_item_changed(self, item: QTableWidgetItem) -> None:
        if self.is_calculating:
            return 

        user_text = item.text()
        self.is_calculating = True 

        if user_text.startswith("="):
            item.setData(Qt.ItemDataRole.UserRole, user_text)
            
            if self.is_formula_view:
                item.setText(user_text)
            else:
                result = self.calculator.parse_and_calculate(user_text, self.table_widget)
                item.setText(result)
        else:
            item.setData(Qt.ItemDataRole.UserRole, None)
            item.setText(user_text) 

        self.is_calculating = False 
        QTimer.singleShot(0, self.recalculate_all_cells)

    def recalculate_all_cells(self) -> None:
        if self.is_calculating: 
            return
            
        self.is_calculating = True
        
        for _ in range(2):
            for r in range(self.table_widget.rowCount()):
                for c in range(self.table_widget.columnCount()):
                    item = self.table_widget.item(r, c)
                    if not item:
                        continue

                    formula = item.data(Qt.ItemDataRole.UserRole)
                    
                    if formula and formula.startswith("="):
                    
                        if self.is_formula_view:
                            if item.text() != formula:
                                item.setText(formula)
                        else:
                            result = self.calculator.parse_and_calculate(formula, self.table_widget)
                            if item.text() != result:
                                item.setText(str(result))

        self.is_calculating = False
        

    #File management
    def new_file(self) -> None:
        try:
            self.current_workbook = Workbook()
            self.active_sheet = self.current_workbook.active
            self.active_sheet.title = "Sheet1"
            self.current_filepath = None 
            default_rows = 10
            default_cols = 5
            
            self.table_widget.blockSignals(True)
            self.table_widget.clear()
            self.table_widget.setRowCount(default_rows)
            self.table_widget.setColumnCount(default_cols)
            self.update_column_headers()
            for r in range(default_rows):
                for c in range(default_cols):
                    self.table_widget.setItem(r, c, QTableWidgetItem(""))
            self.table_widget.blockSignals(False)

            self.action_save.setEnabled(True)
            self.set_edit_actions_enabled(True) 
            self.action_show_formulas.setEnabled(True)
            
            self.setWindowTitle("Редактор - Новий файл*")

        except Exception as e:
            self.table_widget.blockSignals(False)
            QMessageBox.critical(self, "Помилка", f"Не вдалося створити файл: {e}")
            self.reset_app()

    def open_file(self) -> None:
        options = QFileDialog.Options()
        filePath, _ = QFileDialog.getOpenFileName(
            self, "Відкрити Excel файл", "",
            "Excel Files (*.xlsx);;All Files (*)", options=options
        )
        if filePath:
            try:
                self.current_workbook = openpyxl.load_workbook(filePath, data_only=False)
                self.active_sheet = self.current_workbook.active
                self.current_filepath = filePath
                
                self.populate_table() 
                
                self.action_save.setEnabled(True)
                self.set_edit_actions_enabled(True) 
                self.action_show_formulas.setEnabled(True) 
                
                self.setWindowTitle(f"Редактор - {filePath}")
            except Exception as e:
                QMessageBox.critical(self, "Помилка", f"Не вдалося відкрити файл: {e}")
                self.reset_app()


    def populate_table(self) -> None:
        if not self.active_sheet: return
        self.table_widget.blockSignals(True)
        self.table_widget.clear()
        max_row = self.active_sheet.max_row
        max_col = self.active_sheet.max_column
        if max_row == 1 and max_col == 1 and self.active_sheet.cell(1, 1).value is None:
             max_row, max_col = 10, 5
        self.table_widget.setRowCount(max_row)
        self.table_widget.setColumnCount(max_col)
        self.update_column_headers()
        for row_idx, row in enumerate(self.active_sheet.iter_rows(max_row=max_row, max_col=max_col)):
            for col_idx, cell in enumerate(row):
                cell_value = cell.value
                cell_value_str = str(cell_value) if cell_value is not None else ""
                item = QTableWidgetItem()
                if cell_value_str.startswith("="):
                    item.setData(Qt.ItemDataRole.UserRole, cell_value_str)
                    item.setText("...") 
                else:
                    item.setText(cell_value_str)
                self.table_widget.setItem(row_idx, col_idx, item)
        self.table_widget.blockSignals(False)
        self.recalculate_all_cells()
    def save_file(self):
        if not self.current_workbook or not self.active_sheet:
            QMessageBox.warning(self, "Помилка", "Немає активного файлу для збереження.")
            return
        savePath = self.current_filepath
        if savePath is None:
            options = QFileDialog.Options()
            savePath, _ = QFileDialog.getSaveFileName(
                self, "Зберегти новий файл", "Untitled.xlsx", 
                "Excel Files (*.xlsx);;All Files (*)", options=options
            )
        if savePath:
            try:
                for r in range(self.table_widget.rowCount()):
                    for c in range(self.table_widget.columnCount()):
                        item = self.table_widget.item(r, c)
                        value_to_save = None 
                        if item:
                            formula = item.data(Qt.ItemDataRole.UserRole)
                            if formula:
                                value_to_save = formula
                            else:
                                value_to_save = item.text()
                        self.active_sheet.cell(row=r + 1, column=c + 1, value=value_to_save)
                self.current_workbook.save(savePath)
                QMessageBox.information(self, "Успіх", f"Файл успішно збережено у:\n{savePath}")
                self.current_filepath = savePath
                self.setWindowTitle(f"Редактор - {savePath}")
            except Exception as e:
                QMessageBox.critical(self, "Помилка", f"Не вдалося зберегти файл: {e}")

    def reset_app(self) -> None:
        self.current_workbook = None
        self.current_filepath = None
        self.active_sheet = None
        
        self.table_widget.blockSignals(True)
        self.table_widget.clear()
        self.table_widget.setRowCount(0)
        self.table_widget.setColumnCount(0)
        self.table_widget.blockSignals(False)
        
        self.action_save.setEnabled(False)
        self.set_edit_actions_enabled(False)
        
        self.action_show_formulas.setEnabled(False)
        self.action_show_formulas.setChecked(False)
        self.is_formula_view = False
        
        self.setWindowTitle("Редактор Excel (.xlsx) на PySide6")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelEditor()
    window.show()
    sys.exit(app.exec())